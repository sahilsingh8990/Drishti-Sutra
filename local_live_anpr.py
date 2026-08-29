import os
import cv2
import easyocr
import re
import time

from datetime import datetime
from openpyxl import Workbook, load_workbook
from ultralytics import YOLO


# ============================================================
# SETTINGS
# ============================================================

MODEL_PATH = "plate_model.pt"
EXCEL_FILE = "vehicle_data.xlsx"

CAMERA_ID = 0

YOLO_CONFIDENCE = 0.40
OCR_CONFIDENCE = 0.60

DUPLICATE_DELAY = 30

# Number of times same plate must be read before saving
CONFIRMATION_COUNT = 3

# This webcam currently represents Square 1
SQUARE = "Square 1"


# ============================================================
# LOAD YOLO MODEL
# ============================================================

print("Loading plate detector...")

model = YOLO(MODEL_PATH)

print("Plate detector loaded.")


# ============================================================
# LOAD EASYOCR
# ============================================================

print("Loading OCR...")

reader = easyocr.Reader(
    ["en"],
    gpu=False
)

print("OCR loaded.")


# ============================================================
# EXCEL SETUP
# ============================================================

if not os.path.exists(EXCEL_FILE):

    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "ANPR Records"

    sheet.append([
        "Number Plate",
        "Square",
        "Date",
        "Time"
    ])

    workbook.save(EXCEL_FILE)


# ============================================================
# MEMORY
# ============================================================

recent_plates = {}

plate_candidates = {}


# ============================================================
# CLEAN OCR TEXT
# ============================================================

def clean_plate(text):

    text = text.upper()

    text = re.sub(
        r"[^A-Z0-9]",
        "",
        text
    )

    return text

def correct_plate_ocr(plate):

    plate = clean_plate(plate)

    # ==========================================
    # BH / BHARAT SERIES CORRECTION
    #
    # Expected:
    # 25BH2534A
    #
    # EasyOCR may produce:
    # 258H25340
    # ==========================================

    if len(plate) >= 8:

        chars = list(plate)

        # First two characters should be numbers
        number_corrections = {
            "O": "0",
            "Q": "0",
            "I": "1",
            "L": "1",
            "Z": "2",
            "S": "5",
            "B": "8"
        }

        for i in [0, 1]:

            if chars[i] in number_corrections:
                chars[i] = number_corrections[chars[i]]

        # Position 3 should be B
        # OCR commonly reads B as 8

        if chars[2] == "8":
            chars[2] = "B"

        # Position 4 should be H
        # Keep H if correctly recognised

        # BH series numeric section:
        # XXXX

        if (
            chars[2] == "B"
            and chars[3] == "H"
        ):

            for i in range(
                4,
                min(8, len(chars))
            ):

                if chars[i] in number_corrections:
                    chars[i] = number_corrections[chars[i]]

            # Characters AFTER the four-digit number
            # should normally be letters

            letter_corrections = {
                "0": "O",
                "1": "I",
                "2": "Z",
                "5": "S",
                "8": "B"
            }

            for i in range(8, len(chars)):

                if chars[i] in letter_corrections:
                    chars[i] = letter_corrections[chars[i]]

        plate = "".join(chars)

    return plate
# ============================================================
# VALIDATE INDIAN NUMBER PLATE
# ============================================================

def valid_indian_plate(plate):

    patterns = [

        # Examples:
        # MH12AB1234
        # DL01CA1234
        # KA03MN1234

        r"^[A-Z]{2}[0-9]{1,2}[A-Z]{1,3}[0-9]{3,4}$",

        # Bharat Series
        # Example:
        # 22BH6517A

        r"^[0-9]{2}BH[0-9]{4}[A-Z]{1,2}$"
    ]

    for pattern in patterns:

        if re.fullmatch(pattern, plate):
            return True

    return False


# ============================================================
# CONFIRM PLATE ACROSS MULTIPLE FRAMES
# ============================================================

last_candidate = None
candidate_count = 0


def confirm_plate(plate):

    global last_candidate
    global candidate_count

    # New plate detected
    if plate != last_candidate:

        last_candidate = plate
        candidate_count = 1

        print(f"NEW PLATE: {plate}")
        print(f"Candidate: {plate} (1/{CONFIRMATION_COUNT})")

        return False


    # Same plate detected again
    candidate_count += 1

    print(
        f"Candidate: {plate} "
        f"({candidate_count}/{CONFIRMATION_COUNT})"
    )


    # Plate confirmed
    if candidate_count >= CONFIRMATION_COUNT:

        candidate_count = 0
        last_candidate = None

        return True


    return False

# ============================================================
# SAVE TO EXCEL
# ============================================================

def save_plate(plate, square):

    current_time = time.time()

    key = f"{plate}_{square}"


    # --------------------------------------------------------
    # PREVENT DUPLICATES
    # --------------------------------------------------------

    if key in recent_plates:

        elapsed = (
            current_time
            - recent_plates[key]
        )

        if elapsed < DUPLICATE_DELAY:

            print(
                f"Duplicate ignored: {plate}"
            )

            return


    recent_plates[key] = current_time


    # --------------------------------------------------------
    # DATE + TIME
    # --------------------------------------------------------

    now = datetime.now()


    # --------------------------------------------------------
    # SAVE
    # --------------------------------------------------------

    workbook = load_workbook(
        EXCEL_FILE
    )

    sheet = workbook[
        "ANPR Records"
    ]

    sheet.append([
        plate,
        square,
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M:%S")
    ])

    workbook.save(
        EXCEL_FILE
    )


    print()
    print(
        f"SAVED: {plate} | {square}"
    )
    print()


# ============================================================
# OPEN CAMERA
# ============================================================

camera = cv2.VideoCapture(
    CAMERA_ID
)


if not camera.isOpened():

    print(
        "ERROR: Camera could not be opened."
    )

    exit()


# Optional camera resolution

camera.set(
    cv2.CAP_PROP_FRAME_WIDTH,
    1280
)

camera.set(
    cv2.CAP_PROP_FRAME_HEIGHT,
    720
)


print()
print("==============================")
print("LOCAL ANPR STARTED")
print("==============================")
print()
print("Press Q to quit.")
print()


# ============================================================
# MAIN LOOP
# ============================================================

while True:

    success, frame = camera.read()


    if not success:

        print(
            "Could not read camera frame."
        )

        break


    # ========================================================
    # YOLO DETECTION
    # ========================================================

    results = model(
        frame,
        conf=YOLO_CONFIDENCE,
        verbose=False
    )


    # ========================================================
    # PROCESS DETECTIONS
    # ========================================================

    for result in results:

        for box in result.boxes:


            # ------------------------------------------------
            # YOLO CONFIDENCE
            # ------------------------------------------------

            detection_confidence = float(
                box.conf[0]
            )


            # ------------------------------------------------
            # GET BOX
            # ------------------------------------------------

            x1, y1, x2, y2 = map(
                int,
                box.xyxy[0]
            )


            x1 = max(
                0,
                x1
            )

            y1 = max(
                0,
                y1
            )

            x2 = min(
                frame.shape[1],
                x2
            )

            y2 = min(
                frame.shape[0],
                y2
            )


            # ------------------------------------------------
            # CROP PLATE
            # ------------------------------------------------

            plate_crop = frame[
                y1:y2,
                x1:x2
            ]


            if plate_crop.size == 0:
                continue


            # =================================================
            # PREPROCESS FOR OCR
            # =================================================

            plate_crop_large = cv2.resize(
                plate_crop,
                None,
                fx=3,
                fy=3,
                interpolation=cv2.INTER_CUBIC
            )


            gray = cv2.cvtColor(
                plate_crop_large,
                cv2.COLOR_BGR2GRAY
            )


            gray = cv2.bilateralFilter(
                gray,
                11,
                17,
                17
            )


            # Optional contrast improvement

            gray = cv2.equalizeHist(
                gray
            )


            # =================================================
            # OCR
            # =================================================

            ocr_results = reader.readtext(
                gray,
                detail=1,
                paragraph=False
            )


            plate_number = ""

            best_ocr_confidence = 0


            # ------------------------------------------------
            # CHOOSE BEST OCR RESULT
            # ------------------------------------------------

            for detection in ocr_results:

                text = detection[1]

                confidence = detection[2]

                cleaned = correct_plate_ocr(
    text
)


                if confidence > best_ocr_confidence:

                    best_ocr_confidence = confidence

                    plate_number = cleaned


            # =================================================
            # DRAW YOLO BOX
            # =================================================

            cv2.rectangle(
                frame,
                (x1, y1),
                (x2, y2),
                (0, 255, 0),
                2
            )


            # ------------------------------------------------
            # SHOW DETECTION CONFIDENCE
            # ------------------------------------------------

            cv2.putText(
                frame,
                f"Plate {detection_confidence:.2f}",
                (
                    x1,
                    max(
                        25,
                        y1 - 35
                    )
                ),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.6,
                (0, 255, 0),
                2
            )


            # =================================================
            # FILTER OCR RESULT
            # =================================================

            if not plate_number:
                continue


            # Reject very short / long text

            if not (
                7 <= len(plate_number) <= 12
            ):

                continue


            # Reject low OCR confidence

            if (
                best_ocr_confidence
                < OCR_CONFIDENCE
            ):

                continue


            # Reject text that doesn't look like Indian plate

            if not valid_indian_plate(
                plate_number
            ):

                print(
                    f"Rejected OCR: {plate_number}"
                )

                continue


            # =================================================
            # SHOW OCR RESULT
            # =================================================

            cv2.putText(
                frame,
                plate_number,
                (
                    x1,
                    max(
                        50,
                        y1 - 10
                    )
                ),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.8,
                (0, 255, 0),
                2
            )


            print(
                f"Valid OCR: {plate_number} "
                f"| OCR confidence: "
                f"{best_ocr_confidence:.2f}"
            )


            # =================================================
            # CONFIRM PLATE
            # =================================================

            if confirm_plate(
                plate_number
            ):

                print()
                print(
                    f"CONFIRMED: {plate_number}"
                )
                print()


                save_plate(
                    plate_number,
                    SQUARE
                )


    # ========================================================
    # DISPLAY SQUARE
    # ========================================================

    cv2.putText(
        frame,
        SQUARE,
        (20, 40),
        cv2.FONT_HERSHEY_SIMPLEX,
        1,
        (255, 255, 255),
        2
    )


    # ========================================================
    # DISPLAY
    # ========================================================

    cv2.imshow(
        "Local ANPR - Square 1",
        frame
    )


    # ========================================================
    # QUIT
    # ========================================================

    if (
        cv2.waitKey(1)
        & 0xFF
        == ord("q")
    ):

        break


# ============================================================
# CLEANUP
# ============================================================

camera.release()

cv2.destroyAllWindows()

print()
print("ANPR stopped.")