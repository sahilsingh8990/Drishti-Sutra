import cv2
import re
import time
import os
from collections import defaultdict, Counter
from datetime import datetime
from pathlib import Path

import easyocr
import numpy as np
from openpyxl import Workbook, load_workbook
from ultralytics import YOLO

from config import (
    MODEL_PATH,
    OUTPUT_DIR,
    SNAPSHOT_DIR,
    EXCEL_FILE,
    CAMERA_SOURCE,
    CAMERA_WIDTH,
    CAMERA_HEIGHT,
    PLATE_DETECTION_CONFIDENCE,
    YOLO_IMAGE_SIZE,
    OCR_MIN_CONFIDENCE,
    MIN_PLATE_LENGTH,
    MAX_PLATE_LENGTH,
    MIN_CONFIRMATIONS,
    TRACK_TIMEOUT,
    DUPLICATE_COOLDOWN,
    SHOW_OCR_CONFIDENCE,
    SAVE_SNAPSHOT,
    INDIAN_PLATES_ONLY,
)


# ============================================================
# GLOBAL STATE
# ============================================================

# Example:
# observations["MH12AB1234"] = [
#     (timestamp, square, confidence),
#     ...
# ]
observations = defaultdict(list)

# Last time a confirmed plate was written to Excel
last_logged = {}

# EasyOCR
reader = easyocr.Reader(["en"], gpu=False)


# ============================================================
# INITIALIZATION
# ============================================================

def initialize_directories():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)


def initialize_excel():
    if EXCEL_FILE.exists():
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ANPR Records"

    sheet.append([
        "Number Plate",
        "Square",
        "Date",
        "Time",
        "Detection Confidence",
        "OCR Confidence",
        "Snapshot",
    ])

    # Make columns more readable
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 15
    sheet.column_dimensions["C"].width = 15
    sheet.column_dimensions["D"].width = 15
    sheet.column_dimensions["E"].width = 22
    sheet.column_dimensions["F"].width = 18
    sheet.column_dimensions["G"].width = 55

    workbook.save(EXCEL_FILE)


# ============================================================
# SQUARE / REGION DETECTION
# ============================================================

def find_square(center_x, center_y, frame_width, frame_height):
    """
    Divide the frame into four equal quadrants.

        Square 1 | Square 2
        -------------------
        Square 3 | Square 4
    """

    middle_x = frame_width // 2
    middle_y = frame_height // 2

    if center_x < middle_x and center_y < middle_y:
        return "Square 1"

    if center_x >= middle_x and center_y < middle_y:
        return "Square 2"

    if center_x < middle_x and center_y >= middle_y:
        return "Square 3"

    return "Square 4"


def draw_grid(frame):
    height, width = frame.shape[:2]

    middle_x = width // 2
    middle_y = height // 2

    cv2.line(
        frame,
        (middle_x, 0),
        (middle_x, height),
        (255, 255, 255),
        2,
    )

    cv2.line(
        frame,
        (0, middle_y),
        (width, middle_y),
        (255, 255, 255),
        2,
    )

    cv2.putText(
        frame,
        "SQUARE 1",
        (25, 40),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.8,
        (255, 255, 255),
        2,
    )

    cv2.putText(
        frame,
        "SQUARE 2",
        (middle_x + 25, 40),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.8,
        (255, 255, 255),
        2,
    )

    cv2.putText(
        frame,
        "SQUARE 3",
        (25, middle_y + 40),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.8,
        (255, 255, 255),
        2,
    )

    cv2.putText(
        frame,
        "SQUARE 4",
        (middle_x + 25, middle_y + 40),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.8,
        (255, 255, 255),
        2,
    )


# ============================================================
# OCR PREPROCESSING
# ============================================================

def resize_plate(image):
    """
    Upscale very small plates before OCR.
    """

    height, width = image.shape[:2]

    if width < 300:
        scale = 300 / max(width, 1)

        image = cv2.resize(
            image,
            None,
            fx=scale,
            fy=scale,
            interpolation=cv2.INTER_CUBIC,
        )

    return image


def preprocess_plate_variants(plate):
    """
    Generate several versions of the plate.

    OCR is attempted on each one because different lighting
    conditions respond differently to preprocessing.
    """

    if plate is None or plate.size == 0:
        return []

    plate = resize_plate(plate)

    gray = cv2.cvtColor(plate, cv2.COLOR_BGR2GRAY)

    # Mild blur helps remove camera noise
    blurred = cv2.GaussianBlur(gray, (3, 3), 0)

    # CLAHE improves local contrast
    clahe = cv2.createCLAHE(
        clipLimit=2.0,
        tileGridSize=(8, 8),
    )

    contrast = clahe.apply(gray)

    # Otsu threshold
    _, threshold_otsu = cv2.threshold(
        blurred,
        0,
        255,
        cv2.THRESH_BINARY + cv2.THRESH_OTSU,
    )

    # Adaptive threshold
    adaptive = cv2.adaptiveThreshold(
        contrast,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY,
        31,
        11,
    )

    return [
        plate,
        gray,
        contrast,
        threshold_otsu,
        adaptive,
    ]


# ============================================================
# OCR TEXT CLEANING
# ============================================================

def normalize_plate_text(text):
    text = text.upper()

    # Keep only letters and numbers
    text = re.sub(r"[^A-Z0-9]", "", text)

    return text


def valid_generic_plate(text):
    if len(text) < MIN_PLATE_LENGTH:
        return False

    if len(text) > MAX_PLATE_LENGTH:
        return False

    # Require at least one letter and one number
    if not re.search(r"[A-Z]", text):
        return False

    if not re.search(r"[0-9]", text):
        return False

    return True


def valid_indian_plate(text):
    """
    Covers many common Indian vehicle registration patterns.

    Examples:
    MH12AB1234
    DL01CA1234
    KA03MN1234

    It intentionally isn't made overly strict because
    commercial / BH / special registrations can differ.
    """

    patterns = [
        r"^[A-Z]{2}[0-9]{1,2}[A-Z]{1,3}[0-9]{4}$",
        r"^[0-9]{2}BH[0-9]{4}[A-Z]{1,2}$",
    ]

    for pattern in patterns:
        if re.match(pattern, text):
            return True

    return False


def is_valid_plate(text):
    if not valid_generic_plate(text):
        return False

    if INDIAN_PLATES_ONLY:
        return valid_indian_plate(text)

    return True


# ============================================================
# OCR CONFUSION CORRECTION
# ============================================================

def basic_ocr_correction(text):
    """
    Only perform conservative replacements.

    Aggressive O/0, I/1 replacements can actually damage a
    correctly read plate, so validation + temporal voting is
    preferred.
    """

    text = text.replace(" ", "")
    text = text.replace("-", "")

    return text


# ============================================================
# OCR
# ============================================================

def read_plate(plate_crop):
    """
    OCR several preprocessed plate versions and return
    the most likely result.
    """

    variants = preprocess_plate_variants(plate_crop)

    candidates = []

    for variant in variants:
        try:
            results = reader.readtext(
                variant,
                detail=1,
                paragraph=False,
                allowlist="ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789",
            )

        except Exception as error:
            print("OCR error:", error)
            continue

        # Sometimes OCR detects multiple separate chunks,
        # e.g. MH12 + AB + 1234.
        joined_text = ""
        confidence_values = []

        for _, text, confidence in results:
            if confidence < OCR_MIN_CONFIDENCE:
                continue

            cleaned = normalize_plate_text(text)

            if not cleaned:
                continue

            joined_text += cleaned
            confidence_values.append(float(confidence))

        joined_text = basic_ocr_correction(joined_text)

        if not joined_text:
            continue

        if not is_valid_plate(joined_text):
            continue

        if confidence_values:
            average_confidence = (
                sum(confidence_values) /
                len(confidence_values)
            )
        else:
            average_confidence = 0.0

        candidates.append(
            (joined_text, average_confidence)
        )

    if not candidates:
        return None, 0.0

    # First determine text occurring most often
    text_counts = Counter(
        item[0] for item in candidates
    )

    most_common_text = text_counts.most_common(1)[0][0]

    matching_confidences = [
        confidence
        for text, confidence in candidates
        if text == most_common_text
    ]

    best_confidence = max(matching_confidences)

    return most_common_text, best_confidence


# ============================================================
# TEMPORAL RECOGNITION
# ============================================================

def clean_old_observations():
    current_time = time.time()

    expired = []

    for plate_number, records in observations.items():

        filtered = [
            item
            for item in records
            if current_time - item[0] <= TRACK_TIMEOUT
        ]

        observations[plate_number] = filtered

        if not filtered:
            expired.append(plate_number)

    for plate_number in expired:
        observations.pop(plate_number, None)


def add_observation(
    plate_number,
    square,
    ocr_confidence,
):
    observations[plate_number].append(
        (
            time.time(),
            square,
            ocr_confidence,
        )
    )


def plate_is_confirmed(plate_number):
    records = observations.get(
        plate_number,
        [],
    )

    return len(records) >= MIN_CONFIRMATIONS


def confirmed_square(plate_number):
    records = observations.get(
        plate_number,
        [],
    )

    if not records:
        return None

    squares = [
        record[1]
        for record in records
    ]

    return Counter(squares).most_common(1)[0][0]


def confirmed_ocr_confidence(plate_number):
    records = observations.get(
        plate_number,
        [],
    )

    if not records:
        return 0.0

    return max(
        record[2]
        for record in records
    )


# ============================================================
# DUPLICATE PREVENTION
# ============================================================

def duplicate_allowed(plate_number):
    current_time = time.time()

    if plate_number not in last_logged:
        return True

    elapsed = (
        current_time -
        last_logged[plate_number]
    )

    return elapsed >= DUPLICATE_COOLDOWN


# ============================================================
# SNAPSHOT
# ============================================================

def save_snapshot(frame, plate_number, square):
    if not SAVE_SNAPSHOT:
        return ""

    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S_%f"
    )

    safe_square = square.replace(" ", "_")

    filename = (
        f"{plate_number}_"
        f"{safe_square}_"
        f"{timestamp}.jpg"
    )

    path = SNAPSHOT_DIR / filename

    cv2.imwrite(
        str(path),
        frame,
    )

    return str(path)


# ============================================================
# EXCEL
# ============================================================

def log_to_excel(
    plate_number,
    square,
    detection_confidence,
    ocr_confidence,
    snapshot_path,
):
    now = datetime.now()

    try:
        workbook = load_workbook(EXCEL_FILE)

        sheet = workbook["ANPR Records"]

        sheet.append([
            plate_number,
            square,
            now.strftime("%Y-%m-%d"),
            now.strftime("%H:%M:%S"),
            round(detection_confidence, 3),
            round(ocr_confidence, 3),
            snapshot_path,
        ])

        workbook.save(EXCEL_FILE)

        last_logged[plate_number] = time.time()

        print(
            f"[SAVED] "
            f"{plate_number} | "
            f"{square} | "
            f"{now.strftime('%H:%M:%S')}"
        )

    except PermissionError:
        print(
            "\nERROR: Excel file is open."
            "\nClose vehicle_data.xlsx and try again.\n"
        )

    except Exception as error:
        print(
            "Excel saving error:",
            error,
        )


# ============================================================
# DRAW DETECTION
# ============================================================

def draw_plate_detection(
    frame,
    x1,
    y1,
    x2,
    y2,
    plate_number,
    square,
    detection_confidence,
    ocr_confidence,
):
    cv2.rectangle(
        frame,
        (x1, y1),
        (x2, y2),
        (0, 255, 0),
        2,
    )

    if plate_number:

        label = (
            f"{plate_number} | "
            f"{square}"
        )

        if SHOW_OCR_CONFIDENCE:
            label += (
                f" | OCR "
                f"{ocr_confidence:.2f}"
            )

    else:

        label = (
            f"Plate {detection_confidence:.2f}"
        )

    text_y = max(
        25,
        y1 - 10,
    )

    cv2.putText(
        frame,
        label,
        (x1, text_y),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.65,
        (0, 255, 0),
        2,
    )


# ============================================================
# MAIN
# ============================================================

def main():

    initialize_directories()
    initialize_excel()

    # --------------------------------------------------------
    # CHECK MODEL
    # --------------------------------------------------------

    if not MODEL_PATH.exists():

        print("\nERROR")
        print(
            f"License plate model not found:"
            f"\n{MODEL_PATH}"
        )

        print(
            "\nPlace your trained license-plate YOLO model "
            "at:"
        )

        print(
            "plate_model.pt\n"
        )

        return

    # --------------------------------------------------------
    # LOAD MODEL
    # --------------------------------------------------------

    print("Loading license plate detector...")

    plate_model = YOLO(
        str(MODEL_PATH)
    )

    print("License plate detector loaded.")

    # --------------------------------------------------------
    # CAMERA
    # --------------------------------------------------------

    camera = cv2.VideoCapture(
        CAMERA_SOURCE
    )

    camera.set(
        cv2.CAP_PROP_FRAME_WIDTH,
        CAMERA_WIDTH,
    )

    camera.set(
        cv2.CAP_PROP_FRAME_HEIGHT,
        CAMERA_HEIGHT,
    )

    if not camera.isOpened():

        print(
            "ERROR: Camera could not be opened."
        )

        return

    print("\nANPR system started.")
    print("Press Q to quit.\n")

    try:

        while True:

            success, frame = camera.read()

            if not success:
                print(
                    "Could not read camera frame."
                )
                break

            clean_old_observations()

            frame_height, frame_width = (
                frame.shape[:2]
            )

            # ------------------------------------------------
            # YOLO LICENSE PLATE DETECTION
            # ------------------------------------------------

            results = plate_model.predict(
                source=frame,
                conf=PLATE_DETECTION_CONFIDENCE,
                imgsz=YOLO_IMAGE_SIZE,
                verbose=False,
            )

            # Copy used for snapshot before adding all UI
            snapshot_frame = frame.copy()

            # Draw square boundaries
            draw_grid(frame)

            for result in results:

                if result.boxes is None:
                    continue

                for box in result.boxes:

                    detection_confidence = float(
                        box.conf[0].cpu().item()
                    )

                    coordinates = (
                        box.xyxy[0]
                        .cpu()
                        .numpy()
                        .astype(int)
                    )

                    x1, y1, x2, y2 = coordinates

                    # Clamp coordinates
                    x1 = max(
                        0,
                        min(x1, frame_width - 1),
                    )

                    y1 = max(
                        0,
                        min(y1, frame_height - 1),
                    )

                    x2 = max(
                        0,
                        min(x2, frame_width),
                    )

                    y2 = max(
                        0,
                        min(y2, frame_height),
                    )

                    if x2 <= x1 or y2 <= y1:
                        continue

                    # ----------------------------------------
                    # DETERMINE SQUARE
                    # ----------------------------------------

                    center_x = (
                        x1 + x2
                    ) // 2

                    center_y = (
                        y1 + y2
                    ) // 2

                    square = find_square(
                        center_x,
                        center_y,
                        frame_width,
                        frame_height,
                    )

                    # ----------------------------------------
                    # CROP EXACT LICENSE PLATE
                    # ----------------------------------------

                    plate_crop = frame[
                        y1:y2,
                        x1:x2
                    ].copy()

                    # ----------------------------------------
                    # OCR
                    # ----------------------------------------

                    plate_number, ocr_confidence = (
                        read_plate(
                            plate_crop
                        )
                    )

                    if plate_number:

                        add_observation(
                            plate_number,
                            square,
                            ocr_confidence,
                        )

                        # ------------------------------------
                        # CONFIRM ACROSS MULTIPLE FRAMES
                        # ------------------------------------

                        if plate_is_confirmed(
                            plate_number
                        ):

                            stable_square = (
                                confirmed_square(
                                    plate_number
                                )
                            )

                            stable_ocr_conf = (
                                confirmed_ocr_confidence(
                                    plate_number
                                )
                            )

                            # --------------------------------
                            # DUPLICATE CHECK
                            # --------------------------------

                            if duplicate_allowed(
                                plate_number
                            ):

                                # Annotate snapshot
                                cv2.rectangle(
                                    snapshot_frame,
                                    (x1, y1),
                                    (x2, y2),
                                    (0, 255, 0),
                                    3,
                                )

                                cv2.putText(
                                    snapshot_frame,
                                    (
                                        f"{plate_number} | "
                                        f"{stable_square}"
                                    ),
                                    (
                                        x1,
                                        max(25, y1 - 10),
                                    ),
                                    cv2.FONT_HERSHEY_SIMPLEX,
                                    0.8,
                                    (0, 255, 0),
                                    2,
                                )

                                snapshot_path = (
                                    save_snapshot(
                                        snapshot_frame,
                                        plate_number,
                                        stable_square,
                                    )
                                )

                                log_to_excel(
                                    plate_number,
                                    stable_square,
                                    detection_confidence,
                                    stable_ocr_conf,
                                    snapshot_path,
                                )

                    # ----------------------------------------
                    # DISPLAY
                    # ----------------------------------------

                    draw_plate_detection(
                        frame,
                        x1,
                        y1,
                        x2,
                        y2,
                        plate_number,
                        square,
                        detection_confidence,
                        ocr_confidence,
                    )

            # ------------------------------------------------
            # STATUS
            # ------------------------------------------------

            cv2.putText(
                frame,
                "ANPR ACTIVE | Q = Quit",
                (
                    20,
                    frame_height - 20,
                ),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.7,
                (0, 255, 0),
                2,
            )

            cv2.imshow(
                "Automatic Number Plate Recognition",
                frame,
            )

            key = (
                cv2.waitKey(1)
                & 0xFF
            )

            if key == ord("q"):
                break

    except KeyboardInterrupt:
        print("\nStopped by user.")

    finally:

        camera.release()

        cv2.destroyAllWindows()

        print(
            "\nANPR system stopped."
        )


if __name__ == "__main__":
    main()