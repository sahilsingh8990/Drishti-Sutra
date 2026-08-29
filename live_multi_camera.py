import os
import cv2
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook

from inference_sdk import (
    InferenceHTTPClient,
    InferenceConfiguration
)

# ============================================================
# SETTINGS
# ============================================================

EXCEL_FILE = "vehicle_data.xlsx"

PROCESS_INTERVAL = 2
DUPLICATE_DELAY = 30

recent_plates = {}

# ------------------------------------------------------------
# CAMERA SOURCES
# ------------------------------------------------------------

CAMERAS = {
    "Square 1": 0,
}

# For IP cameras, later you can use:
#
# "Square 1": "rtsp://username:password@192.168.1.101/stream"
# "Square 2": "rtsp://username:password@192.168.1.102/stream"


# ============================================================
# ROBOFLOW
# ============================================================

client = InferenceHTTPClient(
    api_url="https://serverless.roboflow.com",
    api_key=os.environ["ROBOFLOW_API_KEY"]
).configure(
    InferenceConfiguration(
        api_key_transport="header"
    )
)


# ============================================================
# EXCEL
# ============================================================

def create_excel():

    if not os.path.exists(EXCEL_FILE):

        workbook = Workbook()

        sheet = workbook.active
        sheet.title = "ANPR Records"

        sheet.append([
            "Number Plate",
            "Square",
            "Date",
            "Time"
        ])

        workbook.save(EXCEL_FILE)


def save_to_excel(plate_number, square):

    current_time = time.time()

    key = f"{plate_number}_{square}"

    if key in recent_plates:

        if current_time - recent_plates[key] < DUPLICATE_DELAY:
            return

    recent_plates[key] = current_time

    now = datetime.now()

    workbook = load_workbook(EXCEL_FILE)

    sheet = workbook["ANPR Records"]

    sheet.append([
        plate_number,
        square,
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M:%S")
    ])

    workbook.save(EXCEL_FILE)

    print(
        f"Saved: {plate_number} | {square}"
    )


# ============================================================
# ROBOFLOW PROCESSING
# ============================================================

def process_frame(frame, square):

    try:

        result = client.run_workflow(
            workspace_name="sahil-singh-nepyt",
            workflow_id="vehicle-number-plate-ocr-1787901794956",
            images={
                "image": frame
            },
            use_cache=False
        )

        if not result:
            return None

        output = result[0]

        plate_texts = output.get(
            "plate_text",
            []
        )

        if not plate_texts:
            return None

        plate_number = plate_texts[0].strip()

        if not plate_number:
            return None

        print(
            f"Detected: {plate_number} | {square}"
        )

        save_to_excel(
            plate_number,
            square
        )

        return plate_number

    except Exception as error:

        print(
            f"Roboflow error for {square}:",
            error
        )

        return None


# ============================================================
# CAMERA INITIALIZATION
# ============================================================

create_excel()

camera_objects = {}

last_processed = {}

latest_plate = {}

for square, source in CAMERAS.items():

    camera = cv2.VideoCapture(source)

    if not camera.isOpened():

        print(
            f"Could not open camera for {square}"
        )

        continue

    camera_objects[square] = camera

    last_processed[square] = 0

    latest_plate[square] = ""

    print(
        f"{square} camera connected."
    )


# ============================================================
# MAIN LOOP
# ============================================================

print("\nMulti-camera ANPR started")
print("Press Q to quit\n")


while True:

    current_time = time.time()

    frames = {}

    for square, camera in camera_objects.items():

        success, frame = camera.read()

        if not success:
            continue

        frame = cv2.resize(
            frame,
            (640, 360)
        )

        frames[square] = frame


        # ----------------------------------------------------
        # PROCESS ROBOFLOW
        # ----------------------------------------------------

        if (
            current_time - last_processed[square]
            >= PROCESS_INTERVAL
        ):

            last_processed[square] = current_time

            plate = process_frame(
                frame.copy(),
                square
            )

            if plate:

                latest_plate[square] = plate


        # ----------------------------------------------------
        # DRAW CAMERA LABEL
        # ----------------------------------------------------

        cv2.putText(
            frame,
            square,
            (20, 35),
            cv2.FONT_HERSHEY_SIMPLEX,
            1,
            (255, 255, 255),
            2
        )


        # ----------------------------------------------------
        # DRAW LAST PLATE
        # ----------------------------------------------------

        if latest_plate[square]:

            cv2.putText(
                frame,
                latest_plate[square],
                (20, 330),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.8,
                (0, 255, 0),
                2
            )


    # ========================================================
    # DISPLAY EACH CAMERA
    # ========================================================

    for square, frame in frames.items():

        cv2.imshow(
            square,
            frame
        )


    if cv2.waitKey(1) & 0xFF == ord("q"):
        break


# ============================================================
# CLEANUP
# ============================================================

for camera in camera_objects.values():

    camera.release()

cv2.destroyAllWindows()

print("ANPR stopped.")